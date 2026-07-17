#!/usr/bin/env python3
"""Generate ByteMail automated test inventory (CSV + XLSX).

Scans test/*_test.dart for test()/testWidgets() cases and emits:
  docs/V1_AUTOMATED_TEST_INVENTORY.csv
  docs/V1_AUTOMATED_TEST_INVENTORY.xlsx
"""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
TEST_DIR = ROOT / "test"
OUT_CSV = ROOT / "docs" / "V1_AUTOMATED_TEST_INVENTORY.csv"
OUT_XLSX = ROOT / "docs" / "V1_AUTOMATED_TEST_INVENTORY.xlsx"
TODAY = date.today().isoformat()

# File-level defaults. Individual rows inherit these; override when needed.
FILE_META: dict[str, dict[str, str]] = {
    "account_service_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0;TA-5",
        "kind": "unit",
        "component": "Auth / Accounts",
        "platform": "All",
    },
    "app_settings_cubit_test.dart": {
        "wave": "W1",
        "tier_refs": "W1;W2;W3;W5",
        "kind": "bloc",
        "component": "Settings",
        "platform": "All",
    },
    "compose_prefill_test.dart": {
        "wave": "W1",
        "tier_refs": "TA-1",
        "kind": "unit",
        "component": "Compose",
        "platform": "All",
    },
    "drift_mail_repository_test.dart": {
        "wave": "W0",
        "tier_refs": "W0;W1;W3;W4",
        "kind": "integration",
        "component": "Data",
        "platform": "All",
    },
    "eml_codec_test.dart": {
        "wave": "W5",
        "tier_refs": "TC-9",
        "kind": "unit",
        "component": "Desktop / EML",
        "platform": "All",
    },
    "focus_scorer_test.dart": {
        "wave": "Foundation",
        "tier_refs": "M6",
        "kind": "unit",
        "component": "Focus",
        "platform": "All",
    },
    "graph_mail_delta_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-1;TC-2",
        "kind": "contract",
        "component": "Sync / Graph",
        "platform": "All",
    },
    "graph_mail_folder_list_test.dart": {
        "wave": "W0",
        "tier_refs": "TC-1",
        "kind": "contract",
        "component": "Sync / Graph",
        "platform": "All",
    },
    "graph_mail_provider_auth_retry_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-5",
        "kind": "contract",
        "component": "Auth / Graph",
        "platform": "All",
    },
    "html_email_fallback_test.dart": {
        "wave": "W5",
        "tier_refs": "DEF-029;DEF-030",
        "kind": "unit",
        "component": "UI / HTML",
        "platform": "Windows",
    },
    "imap_autoconfig_test.dart": {
        "wave": "W0",
        "tier_refs": "TC-8",
        "kind": "unit",
        "component": "Auth / IMAP",
        "platform": "All",
    },
    "imap_folder_role_mapping_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0;TA-1",
        "kind": "unit",
        "component": "Sync / IMAP",
        "platform": "All",
    },
    "imap_list_recent_helpers_test.dart": {
        "wave": "Foundation",
        "tier_refs": "DEF-028",
        "kind": "unit",
        "component": "Sync / IMAP",
        "platform": "All",
    },
    "mail_date_parser_test.dart": {
        "wave": "Foundation",
        "tier_refs": "DEF-027",
        "kind": "unit",
        "component": "Sync / MIME",
        "platform": "All",
    },
    "mail_provider_capabilities_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0",
        "kind": "contract",
        "component": "Protocol",
        "platform": "All",
    },
    "mail_split_layout_test.dart": {
        "wave": "W5",
        "tier_refs": "TB-9",
        "kind": "widget",
        "component": "UI / Desktop",
        "platform": "All",
    },
    "mailbox_cubit_test.dart": {
        "wave": "W1",
        "tier_refs": "W1;W2;TA-1;TA-4",
        "kind": "bloc",
        "component": "Mailbox",
        "platform": "All",
    },
    "mailbox_shortcuts_test.dart": {
        "wave": "W5",
        "tier_refs": "TC-9;DEF-001;DEF-031",
        "kind": "unit",
        "component": "UI / Shortcuts",
        "platform": "Windows",
    },
    "message_actions_repo_test.dart": {
        "wave": "W1",
        "tier_refs": "TA-1;TA-4;W2",
        "kind": "integration",
        "component": "Data / Actions",
        "platform": "All",
    },
    "message_body_find_test.dart": {
        "wave": "W5",
        "tier_refs": "TC-9",
        "kind": "unit",
        "component": "UI / Find",
        "platform": "All",
    },
    "message_body_view_find_test.dart": {
        "wave": "W5",
        "tier_refs": "TC-9",
        "kind": "widget",
        "component": "UI / Find",
        "platform": "All",
    },
    "message_filter_bar_test.dart": {
        "wave": "W2",
        "tier_refs": "TB-2;TB-14",
        "kind": "widget",
        "component": "UI / List",
        "platform": "All",
    },
    "message_headers_sheet_test.dart": {
        "wave": "W1",
        "tier_refs": "TA-1",
        "kind": "unit",
        "component": "UI / Headers",
        "platform": "All",
    },
    "message_list_pane_gestures_test.dart": {
        "wave": "W2",
        "tier_refs": "TB-7;TB-14",
        "kind": "widget",
        "component": "UI / List",
        "platform": "Android",
    },
    "message_list_projector_test.dart": {
        "wave": "W2",
        "tier_refs": "TB-1;TB-2",
        "kind": "unit",
        "component": "UI / List",
        "platform": "All",
    },
    "message_print_service_test.dart": {
        "wave": "W5",
        "tier_refs": "TC-9;DEF-033;DEF-035;DEF-036",
        "kind": "unit",
        "component": "Desktop / Print",
        "platform": "Windows",
    },
    "message_query_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0;TB-2",
        "kind": "unit",
        "component": "Data / Query",
        "platform": "All",
    },
    "mime_builder_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0",
        "kind": "contract",
        "component": "MIME",
        "platform": "All",
    },
    "network_sync_policy_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-2",
        "kind": "unit",
        "component": "Sync / Network",
        "platform": "All",
    },
    "oauth_identity_manager_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-5",
        "kind": "unit",
        "component": "Auth / OAuth",
        "platform": "All",
    },
    "oauth_redirect_capture_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-5",
        "kind": "integration",
        "component": "Auth / OAuth",
        "platform": "Windows",
    },
    "protocol_exception_test.dart": {
        "wave": "Foundation",
        "tier_refs": "DEF-028",
        "kind": "unit",
        "component": "Protocol",
        "platform": "All",
    },
    "provider_dispose_test.dart": {
        "wave": "W3",
        "tier_refs": "TA-0;W3",
        "kind": "unit",
        "component": "Protocol",
        "platform": "All",
    },
    "reading_pane_actions_test.dart": {
        "wave": "W2",
        "tier_refs": "TB-14;TA-1",
        "kind": "widget",
        "component": "UI / Reading",
        "platform": "All",
    },
    "remote_image_policy_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-6",
        "kind": "unit",
        "component": "Privacy / HTML",
        "platform": "All",
    },
    "schema_migration_v4_to_v5_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0",
        "kind": "integration",
        "component": "Data / Schema",
        "platform": "All",
    },
    "schema_v5_test.dart": {
        "wave": "W0",
        "tier_refs": "TA-0",
        "kind": "integration",
        "component": "Data / Schema",
        "platform": "All",
    },
    "send_error_messages_test.dart": {
        "wave": "W4",
        "tier_refs": "TA-2;M4",
        "kind": "unit",
        "component": "Compose / Send",
        "platform": "All",
    },
    "sync_engine_push_wake_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-2",
        "kind": "integration",
        "component": "Sync",
        "platform": "All",
    },
    "sync_engine_send_outbox_test.dart": {
        "wave": "W4",
        "tier_refs": "M4;TA-2",
        "kind": "integration",
        "component": "Sync / Outbox",
        "platform": "All",
    },
    "sync_engine_trash_purge_test.dart": {
        "wave": "W1",
        "tier_refs": "TA-1",
        "kind": "integration",
        "component": "Sync / Trash",
        "platform": "All",
    },
    "sync_profile_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-1;TB-10",
        "kind": "integration",
        "component": "Sync / Profiles",
        "platform": "All",
    },
    "sync_status_sheet_test.dart": {
        "wave": "W3",
        "tier_refs": "TC-5",
        "kind": "widget",
        "component": "UI / Sync",
        "platform": "All",
    },
    "theme_tokens_test.dart": {
        "wave": "W7",
        "tier_refs": "M8;UI-L8;DEF-019",
        "kind": "unit",
        "component": "UI / Theme",
        "platform": "All",
    },
    "thread_id_test.dart": {
        "wave": "W2",
        "tier_refs": "TB-1",
        "kind": "unit",
        "component": "Data / Threads",
        "platform": "All",
    },
    "widget_test.dart": {
        "wave": "Foundation",
        "tier_refs": "M0",
        "kind": "smoke",
        "component": "UI / Shell",
        "platform": "All",
    },
}

# Suite-level wave overrides from Renee inventory (file, group) -> wave.
GROUP_WAVE: dict[tuple[str, str], str] = {
    ("app_settings_cubit_test.dart", "AppSettingsCubit trashRetentionDays"): "W1",
    ("app_settings_cubit_test.dart", "AppSettingsCubit threadDisplayMode"): "W2",
    ("app_settings_cubit_test.dart", "AppSettingsCubit swipe actions"): "W2",
    ("app_settings_cubit_test.dart", "AppSettingsCubit blockRemoteImages"): "W3",
    ("app_settings_cubit_test.dart", "AppSettingsCubit pushOnCellular"): "W3",
    ("app_settings_cubit_test.dart", "AppSettingsCubit readingPanePosition"): "W5",
    ("app_settings_cubit_test.dart", "AppSettingsCubit visualFocusEnabled"): "W5",
    ("drift_mail_repository_test.dart", "DriftMailRepository.wipeAccount"): "W0",
    ("drift_mail_repository_test.dart", "DriftMailRepository.setUnreadBulk"): "W1",
    ("drift_mail_repository_test.dart", "DriftMailRepository.recountUnreadCounts"): "W2",
    ("drift_mail_repository_test.dart", "DriftMailRepository.updateMessageRawHeaders"): "W1",
    ("drift_mail_repository_test.dart", "DriftMailRepository outbox recipients"): "W4",
    ("drift_mail_repository_test.dart", "DriftMailRepository sync job viewer"): "W3",
    ("mailbox_cubit_test.dart", "MailboxCubit mark read/unread"): "W1",
    ("mailbox_cubit_test.dart", "MailboxCubit ensureHeadersCached"): "W1",
    ("mailbox_cubit_test.dart", "MailboxCubit ensureSystemFolder"): "W1",
    ("mailbox_cubit_test.dart", "MailboxCubit message actions"): "W1",
    ("message_actions_repo_test.dart", "DriftMailRepository.resolveFolderByRole"): "W1",
    ("message_actions_repo_test.dart", "DriftMailRepository.moveMessageLocal"): "W1",
    ("message_actions_repo_test.dart", "DriftMailRepository.hardDeleteLocal"): "W1",
    ("message_actions_repo_test.dart", "DriftMailRepository.listTrashedPastRetention"): "W1",
    ("message_actions_repo_test.dart", "DriftMailRepository pin and snooze"): "W2",
    ("message_query_test.dart", "MessageQuery.matches predicate stacking"): "W0",
    ("message_query_test.dart", "DriftMailRepository MessageQuery SQL"): "W0",
    ("message_query_test.dart", "upsertMessages merge policy"): "W0",
    ("sync_engine_send_outbox_test.dart", "splitOutboxRecipients"): "W4",
    ("sync_engine_send_outbox_test.dart", "SyncEngine send_outbox"): "W4",
    ("sync_profile_test.dart", "ResolvedSyncPolicy.allowsFolder"): "W3",
    ("sync_profile_test.dart", "SyncProfile.resolvePolicy"): "W3",
    ("sync_profile_test.dart", "applyRetention account scope"): "W3",
    ("sync_profile_test.dart", "SyncEngine folder scope"): "W3",
    ("sync_profile_test.dart", "MessageBodyCache bodyPolicy"): "W3",
}

# Case-level wave overrides (file, test_name) -> wave.
CASE_WAVE: dict[tuple[str, str], str] = {
    ("mailbox_cubit_test.dart", "togglePinSelected pins locally without sync jobs"): "W2",
    ("mailbox_cubit_test.dart", "snoozeSelected hides message until expiry then refresh shows it"): "W2",
    ("mailbox_cubit_test.dart", "markFocusBucket upserts sender rule and updates message"): "Foundation",
    ("mailbox_cubit_test.dart", "markFocusBucket domain scope upserts domain rule"): "Foundation",
}

# Case-level kind overrides (file, test_name) -> kind.
CASE_KIND: dict[tuple[str, str], str] = {
    ("mailbox_shortcuts_test.dart", "isEditingText ignores readOnly SelectableText"): "widget",
    ("mailbox_shortcuts_test.dart", "isEditingText is true for focused TextField"): "widget",
    ("mailbox_shortcuts_test.dart", "keymap help sheet lists bindings"): "widget",
    ("message_query_test.dart", "defaults exclude future snooze, drafts, and trash"): "unit",
    ("message_query_test.dart", "defaults hide snoozed, draft, and trashed rows"): "integration",
    ("message_query_test.dart", "starredOnly returns only starred non-excluded rows"): "integration",
    ("message_query_test.dart", "pinnedOnly and snoozedOnly SQL filters"): "integration",
    ("message_query_test.dart", "userFilter unread, sender, date, attachments, keyword via FTS"): "integration",
    ("message_query_test.dart", "includeDrafts and includeTrashed surface excluded rows"): "integration",
    ("message_query_test.dart", "preserves existing threadId when incoming threadId is null"): "integration",
    ("message_query_test.dart", "DEF-007: keeps local read when sync would re-unread"): "integration",
    ("message_query_test.dart", "allows remote mark-read when local is still unread"): "integration",
    ("sync_engine_send_outbox_test.dart", "splits comma and semicolon lists"): "unit",
    ("sync_engine_send_outbox_test.dart", "decodes JSON address arrays"): "unit",
    ("sync_status_sheet_test.dart", "SyncJob.errorSnippet reads failed cursorJson"): "unit",
    ("sync_status_sheet_test.dart", "SyncStatusSheetBody shows jobs and retry"): "widget",
    ("reading_pane_actions_test.dart", "wide breakpoint constant is 520"): "unit",
    ("sync_profile_test.dart", "null scope allows all; roles and remote ids match"): "unit",
}

# Case-level platform overrides.
CASE_PLATFORM: dict[tuple[str, str], str] = {
    ("message_list_pane_gestures_test.dart", "swipe is not wrapped on non-Android platforms"): "Non-Android",
}

# Group-level kind overrides (file, group) -> kind.
GROUP_KIND: dict[tuple[str, str], str] = {
    ("message_query_test.dart", "MessageQuery.matches predicate stacking"): "unit",
    ("message_query_test.dart", "DriftMailRepository MessageQuery SQL"): "integration",
    ("message_query_test.dart", "upsertMessages merge policy"): "integration",
    ("sync_profile_test.dart", "ResolvedSyncPolicy.allowsFolder"): "unit",
    ("sync_profile_test.dart", "SyncProfile.resolvePolicy"): "integration",
    ("sync_profile_test.dart", "applyRetention account scope"): "integration",
    ("sync_profile_test.dart", "SyncEngine folder scope"): "integration",
    ("sync_profile_test.dart", "MessageBodyCache bodyPolicy"): "integration",
    ("sync_engine_send_outbox_test.dart", "splitOutboxRecipients"): "unit",
    ("sync_engine_send_outbox_test.dart", "SyncEngine send_outbox"): "integration",
    ("reading_pane_actions_test.dart", "ReadingPane adaptive actions"): "widget",
}

COLUMNS = [
    "test_id",
    "wave",
    "tier_refs",
    "test_file",
    "test_group",
    "test_name",
    "kind",
    "component",
    "platform",
    "asserts",
    "status",
    "evaluation_status",
    "added_date",
    "last_verified_wave",
    "manual_companion",
    "notes",
]


@dataclass
class Case:
    file_name: str
    group: str
    name: str
    kind_hint: str
    line: int


STRING_RE = re.compile(
    r"""(?P<kw>testWidgets|test)\s*\(\s*(?:'(?P<sname>(?:\\'|[^'])*)'|"(?P<dname>(?:\\"|[^"])*)")""",
    re.MULTILINE,
)
GROUP_RE = re.compile(
    r"""group\s*\(\s*(?:'(?P<sname>(?:\\'|[^'])*)'|"(?P<dname>(?:\\"|[^"])*)")""",
    re.MULTILINE,
)


def infer_asserts(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", name.strip())
    if cleaned.endswith("."):
        return cleaned
    return cleaned[0].upper() + cleaned[1:] if cleaned else cleaned


def parse_file(path: Path) -> list[Case]:
    text = path.read_text(encoding="utf-8")
    events: list[tuple[int, str, str, str]] = []
    def _captured_name(match: re.Match[str]) -> str:
        return match.group("sname") if match.group("sname") is not None else match.group("dname")

    for m in GROUP_RE.finditer(text):
        events.append((m.start(), "group", _captured_name(m), ""))
    for m in STRING_RE.finditer(text):
        events.append((m.start(), "case", _captured_name(m), m.group("kw")))
    events.sort(key=lambda e: e[0])

    stack: list[tuple[str, int]] = []  # (group_name, open_brace_index approx)
    # Track nesting via brace depth from each group start.
    cases: list[Case] = []
    # Simpler approach: maintain current group by scanning with brace depth.
    depth = 0
    group_stack: list[tuple[int, str]] = []  # depth_at_group_body, name
    i = 0
    pending_group: str | None = None
    pending_case: tuple[str, str] | None = None
    # Rebuild with a linear scan using event positions + braces between events.
    pos = 0
    current_group = "—"
    group_depth_stack: list[tuple[int, str]] = []

    # Brace-aware walk using events interleaved with source.
    for start, kind, name, kw in events:
        # Count braces from pos to start
        segment = text[pos:start]
        for ch in segment:
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                while group_depth_stack and depth < group_depth_stack[-1][0]:
                    group_depth_stack.pop()
                current_group = group_depth_stack[-1][1] if group_depth_stack else "—"
        pos = start
        if kind == "group":
            # Find the opening '{' after this group call to set body depth.
            brace = text.find("{", start)
            body_depth = depth + 1  # will be inside that brace once entered
            # Advance pos through to brace inclusively for depth accounting later? Keep event-based.
            group_depth_stack.append((body_depth, name))
            current_group = name
        else:
            line = text.count("\n", 0, start) + 1
            cases.append(
                Case(
                    file_name=path.name,
                    group=current_group,
                    name=name.replace("\\'", "'").replace('\\"', '"'),
                    kind_hint="widget" if kw == "testWidgets" else "unit",
                    line=line,
                )
            )
    return cases


def wave_sort_key(wave: str) -> tuple[int, str]:
    order = {
        "Foundation": 0,
        "W0": 1,
        "W1": 2,
        "W2": 3,
        "W3": 4,
        "W4": 5,
        "W5": 6,
        "W6": 7,
        "W7": 8,
        "X": 9,
        "Unmapped": 10,
    }
    return (order.get(wave, 99), wave)


def manual_companion(wave: str, file_name: str) -> str:
    if wave == "W5" or file_name in {
        "mail_split_layout_test.dart",
        "mailbox_shortcuts_test.dart",
        "message_print_service_test.dart",
        "message_body_find_test.dart",
        "message_body_view_find_test.dart",
        "eml_codec_test.dart",
        "html_email_fallback_test.dart",
    }:
        return "W5_WINDOWS_CHECKLIST"
    if wave == "W2" or file_name in {
        "message_list_pane_gestures_test.dart",
        "message_filter_bar_test.dart",
    }:
        return "W2_AVD_CHECKLIST"
    return "—"


def build_rows() -> list[dict[str, str]]:
    files = sorted({p.resolve() for p in TEST_DIR.glob("*_test.dart")}, key=lambda p: p.name.lower())
    rows: list[dict[str, str]] = []
    counters: dict[str, int] = defaultdict(int)

    for path in files:
        meta = FILE_META.get(
            path.name,
            {
                "wave": "Unmapped",
                "tier_refs": "—",
                "kind": "unit",
                "component": "Unmapped",
                "platform": "All",
            },
        )
        wave = meta["wave"]
        cases = parse_file(path)
        if not cases:
            counters[wave] += 1
            tid = f"AT-{wave}-{counters[wave]:03d}"
            rows.append(
                {
                    "test_id": tid,
                    "wave": wave,
                    "tier_refs": meta["tier_refs"],
                    "test_file": f"test/{path.name}",
                    "test_group": "—",
                    "test_name": "*",
                    "kind": meta["kind"],
                    "component": meta["component"],
                    "platform": meta["platform"],
                    "asserts": "File-level coverage (no parseable named cases)",
                    "status": "active",
                    "evaluation_status": "Cataloged",
                    "added_date": TODAY,
                    "last_verified_wave": wave if wave not in {"X", "Unmapped", "Foundation"} else "—",
                    "manual_companion": manual_companion(wave, path.name),
                    "notes": "Parser found no test()/testWidgets() string names",
                }
            )
            continue

        for case in cases:
            case_wave = (
                CASE_WAVE.get((path.name, case.name))
                or GROUP_WAVE.get((path.name, case.group))
                or meta["wave"]
            )
            kind = (
                CASE_KIND.get((path.name, case.name))
                or GROUP_KIND.get((path.name, case.group))
                or meta["kind"]
            )
            if case.kind_hint == "widget" and kind in {"unit", "bloc"}:
                # Prefer explicit widget harness when file default is non-widget.
                if (path.name, case.name) not in CASE_KIND:
                    kind = "widget"
            platform = CASE_PLATFORM.get((path.name, case.name), meta["platform"])
            rows.append(
                {
                    "test_id": "PENDING",
                    "wave": case_wave,
                    "tier_refs": meta["tier_refs"],
                    "test_file": f"test/{path.name}",
                    "test_group": case.group,
                    "test_name": case.name,
                    "kind": kind,
                    "component": meta["component"],
                    "platform": platform,
                    "asserts": infer_asserts(case.name),
                    "status": "active",
                    "evaluation_status": "Cataloged",
                    "added_date": TODAY,
                    "last_verified_wave": (
                        case_wave if case_wave not in {"X", "Unmapped", "Foundation"} else "—"
                    ),
                    "manual_companion": manual_companion(case_wave, path.name),
                    "notes": f"L{case.line}",
                }
            )

    rows.sort(key=lambda r: (wave_sort_key(r["wave"]), r["test_file"], r["test_group"], r["test_name"]))
    # Reassign stable IDs after sort within wave.
    counters.clear()
    for row in rows:
        wave = row["wave"]
        counters[wave] += 1
        row["test_id"] = f"AT-{wave}-{counters[wave]:03d}"
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        values = []
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            values.append(len(str(cell.value)))
        width = min(max(values or [10]) + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()

    # Automated sheet
    ws = wb.active
    ws.title = "Automated"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row[c] for c in COLUMNS])
    style_header(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    ws.freeze_panes = "A2"
    table = Table(displayName="AutomatedTests", ref=f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    autosize(ws)

    # By_File
    by_file = Counter(r["test_file"] for r in rows)
    wf = wb.create_sheet("By_File")
    wf.append(["test_file", "case_count", "wave", "kind", "component", "platform"])
    file_meta = {}
    for r in rows:
        file_meta.setdefault(
            r["test_file"],
            {
                "wave": r["wave"],
                "kind": r["kind"],
                "component": r["component"],
                "platform": r["platform"],
            },
        )
    for path, count in sorted(by_file.items()):
        meta = file_meta[path]
        wf.append([path, count, meta["wave"], meta["kind"], meta["component"], meta["platform"]])
    style_header(wf)
    wf.auto_filter.ref = f"A1:F{len(by_file) + 1}"
    wf.freeze_panes = "A2"
    autosize(wf)

    # Wave_Summary
    ws_sum = wb.create_sheet("Wave_Summary")
    ws_sum.append(["wave", "case_count", "file_count", "unit", "widget", "bloc", "integration", "contract", "smoke", "notes"])
    by_wave_cases = Counter(r["wave"] for r in rows)
    by_wave_files: dict[str, set[str]] = defaultdict(set)
    by_wave_kind: dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        by_wave_files[r["wave"]].add(r["test_file"])
        by_wave_kind[r["wave"]][r["kind"]] += 1
    for wave, count in sorted(by_wave_cases.items(), key=lambda kv: wave_sort_key(kv[0])):
        kinds = by_wave_kind[wave]
        note = {
            "Foundation": "Pre-wave / milestone foundation",
            "X": "Cross-cutting (touches multiple waves)",
            "Unmapped": "Needs wave assignment",
            "W5": "In progress — inventory refreshed from Renee suite overrides",
            "W4": "Compose wave mostly pending; early outbox/send coverage present",
            "W7": "Theme polish / hardening samples",
        }.get(wave, "")
        ws_sum.append(
            [
                wave,
                count,
                len(by_wave_files[wave]),
                kinds.get("unit", 0),
                kinds.get("widget", 0),
                kinds.get("bloc", 0),
                kinds.get("integration", 0),
                kinds.get("contract", 0),
                kinds.get("smoke", 0),
                note,
            ]
        )
    style_header(ws_sum)
    autosize(ws_sum)

    # Coverage_Gaps
    wg = wb.create_sheet("Coverage_Gaps")
    wg.append(["area", "gap", "severity", "suggested_owner", "notes"])
    gaps = [
        (
            "integration_test/",
            "No Flutter integration_test/ suite",
            "Pri-2",
            "Renee",
            "Device/UI flows are manual checklists today",
        ),
        (
            "Android native",
            "No androidTest / JVM widget provider tests",
            "Pri-3",
            "Renee",
            "Kotlin AppWidgetProvider untested automatically",
        ),
        (
            "Windows native",
            "No C++/WinRT automated tests",
            "Pri-3",
            "Renee",
            "Tray/print HWND behavior is manual / unit-adjacent",
        ),
        (
            "W4 Compose",
            "Full compose/attachments/signatures wave not landed",
            "Pri-1",
            "Jules+Renee",
            "compose_prefill exists; W4 E2E pending",
        ),
        (
            "W6 Notifications",
            "No notification service automated tests yet",
            "Pri-1",
            "Renee",
            "Blocked until W5 lands",
        ),
        (
            "CI",
            "No GitHub Actions workflow for flutter test",
            "Pri-2",
            "Steve",
            "Inventory assumes local/agent flutter test runs",
        ),
        (
            "Manual E2E matrix",
            "docs/V1_MANUAL_E2E_MATRIX.csv not created yet",
            "Pri-2",
            "Page",
            "ROADMAP FW-5; keep separate from this inventory",
        ),
        (
            "Runtime results",
            "evaluation_status is Cataloged, not Pass/Fail",
            "Pri-3",
            "Renee",
            "Do not treat catalog rows as verified green without a run",
        ),
    ]
    for gap in gaps:
        wg.append(list(gap))
    style_header(wg)
    autosize(wg, max_width=60)

    # Readme
    wr = wb.create_sheet("Readme")
    wr.append(["Field", "Value"])
    info = [
        ("Canonical git artifact", "docs/V1_AUTOMATED_TEST_INVENTORY.csv"),
        ("Operator workbook", "docs/V1_AUTOMATED_TEST_INVENTORY.xlsx"),
        ("Generator", "tool/generate_test_inventory.py"),
        ("Generated", TODAY),
        ("Total cases", str(len(rows))),
        ("Total files", str(len(by_file))),
        ("Update ritual", "After each wave land: Renee emits delta → Page regenerates/updates inventory → Steve gates land"),
        ("evaluation_status", "Cataloged = present in suite; Pass/Fail only after an explicit flutter test run"),
        ("Manual companion", "Separate checklists / future V1_MANUAL_E2E_MATRIX.csv"),
    ]
    for row in info:
        wr.append(list(row))
    style_header(wr)
    autosize(wr, max_width=80)

    wb.save(OUT_XLSX)


def main() -> None:
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    by_wave = Counter(r["wave"] for r in rows)
    by_file = Counter(r["test_file"] for r in rows)
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_XLSX}")
    print(f"cases={len(rows)} files={len(by_file)}")
    for wave, count in sorted(by_wave.items(), key=lambda kv: wave_sort_key(kv[0])):
        print(f"  {wave}: {count}")


if __name__ == "__main__":
    main()
